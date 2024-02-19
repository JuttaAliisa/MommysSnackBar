import gspread
from google.oauth2.service_account import Credentials
from os import system, name
from colorama import Fore, Back, Style
from openpyxl import load_workbook

SCOPE = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/drive"
    ]

CREDS = Credentials.from_service_account_file('creds.json')
SCOPED_CREDS = CREDS.with_scopes(SCOPE)
GSPREAD_CLIENT = gspread.authorize(SCOPED_CREDS)
SPREADSHEET_NAME = 'mommys_snackbar_helper'

def clear():
    '''
    Clears terminal
    '''
    system('cls' if name == 'nt' else 'clear')

def consumption(worksheet):
    """
    function for inputting consumption values
    """
    #clear terminal for better readability
    clear()
    """
    # Select the appropriate worksheet
    worksheet = SHEET.worksheet('usage')
    stock_sheet = SHEET.worksheet('stock')
    """

    # Get user input
    print(Fore.MAGENTA)
    print(Style.BRIGHT + "Snack Consumption logging \n")
    print(Style.RESET_ALL)
    print(Fore.BLUE)
    print('When taking a snack from the stock, please insert correct number for the snack')
    print(Style.RESET_ALL)
    
    products = worksheet.row_values(1)
    numbers = [1, 2, 3, 4, 5, 6]
    print(Fore.BLUE)
    for product, number in zip(products, numbers):
        print(f"{product} = {number}")
    print(Style.RESET_ALL)

    user_choice = input("What did you take? Enter number: ")

    # Define the values based on user input and validating input
    if user_choice == '1':
        values_to_add = [1, 0, 0, 0, 0, 0]
        cell_value = worksheet.cell(1, 1).value
    elif user_choice == '2':
        values_to_add = [0, 1, 0, 0, 0, 0]
        cell_value = worksheet.cell(1, 2).value
    elif user_choice == '3':
        values_to_add = [0, 0, 1, 0, 0, 0]
        cell_value = worksheet.cell(1, 3).value
    elif user_choice == '4':
        values_to_add = [0, 0, 0, 1, 0, 0]
        cell_value = worksheet.cell(1, 4).value
    elif user_choice == '5':
        values_to_add = [0, 0, 0, 0, 1, 0]
        cell_value = worksheet.cell(1, 5).value
    elif user_choice == '6':
        values_to_add = [0, 0, 0, 0, 0, 1]
        cell_value = worksheet.cell(1, 6).value
    else:
        print(Fore.RED)
        print("\nInvalid choice. Please enter a number from 1 through 6 \n")
        print(Style.RESET_ALL)
        print("Press enter to try again")
        input()
        consumption(spreadsheet.worksheet('usage'))


    # Add the values to the next available row
    worksheet.append_row(values_to_add)
    # Turn the value negative and add to stock sheet
    negative_values = [-x for x in values_to_add]
    worksheet.append_row(negative_values)

    print(f"\nThank you! 1pc of {cell_value} deleted from SnackBar stock \n")
    print('Do you want to take out another item? \n')

    def get_user_choice():
        user_choice = input("y=yes, n=no: ")
        return user_choice.lower()


    while True:
        user_choice = get_user_choice()

        if user_choice == 'y':
            consumption()
            break  # exit the loop after consumption
        elif user_choice == 'n':
            print("Thank you for your contribution!\n")
            print("Press enter to get back to start\n")
            input("")
            start()
            break  # exit the loop after starting again
        else:
            print(Fore.RED)
            print("Invalid choice. Please enter y or n: ")
            print(Style.RESET_ALL)

def stock(worksheet):
    """
    Function to check current stock value
    """
    #clear terminal for better readability
    clear()
    print(Fore.MAGENTA)
    print(Style.BRIGHT + "The current stock is: \n")
    print(Style.RESET_ALL)
    print(Fore.GREEN)
    #print user friendly stock values
    products = worksheet.row_values(1)
    amounts = worksheet.row_values(2)
    for product, amount in zip(products, amounts):
        print(f"{product}: {amount}pc")
    print(Style.RESET_ALL)

    print("\nPress enter to go back to main page")
    input()
    start()
        
def recommendation(usage_sheet, stock_sheet, recommendation_sheet):
    """
    Function to calculate purchase recommendation
    """
    #clear terminal for better readability
    clear()
    print(Fore.MAGENTA)
    print(Style.BRIGHT + "Restock recommendations for a shopping trip \n")
    print(Style.RESET_ALL)
    print("Please consider purchasing the following snacks to make sure the stock lasts:")
    print(Fore.GREEN)
    
    # Get the values from the "usage" and "stock" sheets
    usage_values = usage_sheet.row_values(2)
    stock_values = stock_sheet.row_values(2)
    
    # Convert the values to integers
    usage_values = [int(value) for value in usage_values]
    stock_values = [int(value) for value in stock_values]

    # Calculate the difference between usage and stock
    # If stock has gone to 0, we will add one to cover increasing demand
    recommendation_values = [usage - stock + 1 if stock == 0 else usage - stock for usage, stock in zip(usage_values, stock_values)]

    # Add the values to the next available row
    recommendation_sheet.update([recommendation_values], 'A2')

    #print user friendly recommendation list
    products = recommendation_sheet.row_values(1)
    amounts = recommendation_values
    for product, amount in zip(products, amounts):
        if amount > 0:
            print(f"{product}: {amount}pc")
        else:
            print(f"{product} has enough stock")
    print(Style.RESET_ALL)
    print("\nPress enter to go back to main page")
    input()
    start()

def restock():
    """
    Function to perform restock of the snack bar
    """
    #clear terminal for better readability
    clear()
    print(Fore.MAGENTA)
    print(Style.BRIGHT + "Welcome to restock feature!")
    print(Style.RESET_ALL)
    print(Fore.BLUE)
    print("Please add stocked amount for every item")
    print(Style.RESET_ALL)
    print(Fore.GREEN)

    # Select the appropriate worksheet
    restock_sheet = SHEET.worksheet('restock')
    usage_sheet = SHEET.worksheet('usage')
    stock_sheet = SHEET.worksheet('stock')

    # Get the header row as a list of Cell objects
    header_row = restock_sheet.row_values(1)
    header_cells = restock_sheet.range(2, 1, 2, len(header_row))

    # List to store restock quantities
    restock_values = []

    # Get all values in the usage sheet
    all_usage_values = usage_sheet.get_all_values()

    # Get restock amounts from the user with input validation
    for i, product in enumerate(header_row):
        while True:
            try:
                quantity = int(input(f"How many {product}s restocked: "))
                # Check if the input is non-negative
                if quantity >= 0:
                    break
                else:
                    print("Please enter a non-negative number.")
            except ValueError:
                print("Invalid input. Please enter a valid number.")

        # Check if the input is non-zero before updating the cells
        if quantity != 0:
            header_cells[i].value = quantity
            for j in range(2, len(all_usage_values)):
                if all_usage_values[j][i].replace('.', '', 1).isdigit():
                    usage_sheet.update_cell(j + 1, i + 1, '')
        
        restock_values.append(quantity)

    # Update the cells with the new values
    restock_sheet.update_cells(header_cells)
    stock_sheet.append_row(restock_values)

    print(Style.RESET_ALL)
    print("Restock logged. Thank you for restocking!\n")
    print("Press enter to go back to main page")
    input()
    start()

def start():
    clear()
    print("Welcome to Mommy's SnackBar Stock Helper")
    print("Help Mom and log your activity around snack bar.")
    print("This way we can make sure none of is left with out a snack aver again!")
    print("")
    print("Would you like to:")
    print("1: Take a snack")
    print("2: Restock snacks")
    print("3: Check current stock")
    print("4: Get shopping recommendations")
    print("5: Stocktaking")
    user_choice = input("Enter number: ")

    spreadsheet = GSPREAD_CLIENT.open(SPREADSHEET_NAME)

    # Define the values based on user input
    if user_choice == '1':
        consumption(spreadsheet.worksheet('usage'))
    elif user_choice == '2':
        restock()
    elif user_choice == '3':
        stock(spreadsheet.worksheet('stock'))
    elif user_choice == '4':
        recommendation(spreadsheet.worksheet('usage'), spreadsheet.worksheet('stock'), spreadsheet.worksheet('recommendation'))
    elif user_choice == '5':
        stocktaking()
    else:
        print("Invalid choice. Please enter a number between 1 and 4")
        exit()

def stocktaking():
    #clear terminal for better readability
    clear()

    print("Welcome to stocktaking\n")
    print("If you feel or know that your stock is not accurate, stocktaking will help you")
    print("Please note that if you proceed with stocktaking, your stock history will be deleted")
    print("This will affect the recommendations temporarily")
    print("Want to proceed with stock taking?")

    def get_user_choice():
        user_choice = input("y=yes, n=no: ")
        return user_choice.lower()


    while True:
        user_choice = get_user_choice()

        if user_choice == 'y':
            # Select the appropriate worksheets
            usage_sheet = SHEET.worksheet('usage')
            stock_sheet = SHEET.worksheet('stock')
            restock_sheet = SHEET.worksheet('restock')

            # Define the ranges to clear
            usage_range = 'A3:ZZ'
            stock_range = 'A3:ZZ'
            restock_range = 'A2:ZZ'

            # Clear everything from the usage sheet starting at row 3
            usage_sheet.batch_clear([usage_range])

            # Clear everything from the stock sheet starting at row 3
            stock_sheet.batch_clear([stock_range])

            # Clear everything from the restock sheet starting at row 2
            restock_sheet.batch_clear([restock_range])

            print("Stockdata has been cleared.\n Please proceed to restocking the actual stock you have by pressing enter")
            input()
            restock()
            break  # exit the loop
        elif user_choice == 'n':
            print("Stocktaking cancelled")
            print("Press enter to get back to start\n")
            input("")
            start()
            break  # exit the loop
        else:
            print(Fore.RED)
            print("Invalid choice. Please enter y or n: ")
            print(Style.RESET_ALL)
    
    def stocktake():
        # Select the appropriate worksheets
        usage_sheet = SHEET.worksheet('usage')
        stock_sheet = SHEET.worksheet('stock')
        restock_sheet = SHEET.worksheet('restock')

        # Define the ranges to clear
        usage_range = 'A3:ZZ'
        stock_range = 'A3:ZZ'
        restock_range = 'A2:ZZ'

        # Clear everything from the usage sheet starting at row 3
        usage_sheet.batch_clear([usage_range])

        # Clear everything from the stock sheet starting at row 3
        stock_sheet.batch_clear([stock_range])

        # Clear everything from the restock sheet starting at row 2
        restock_sheet.batch_clear([restock_range])

        print("Stockdata has been cleared.\n Please proceed to restocking the actual stock you have by pressing enter")
        input()
        restock()

start()
